import openpyxl


### 读
from openpyxl.cell import Cell

wb=openpyxl.load_workbook(r"/opt/test_excel.xlsx")
sheet = wb.worksheets[0]

for row in sheet.rows:
    for cell in row:
        # print(cell.coordinate, cell.value, end=", ")
        print(cell.value, end=", ")
    print()




# ## 写
#
# wb = openpyxl.Workbook()
#
# # grab the active worksheet
# ws = wb.active#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
#
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# wb.save(r"/opt/test_excel.xlsx")

